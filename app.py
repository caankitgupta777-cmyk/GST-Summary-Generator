import os
import tempfile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, jsonify, send_file
from parser import process_file

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = tempfile.gettempdir()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/upload', methods=['POST'])
def upload_files():
    if 'files[]' not in request.files:
        return jsonify({'error': 'No files provided'}), 400
        
    files = request.files.getlist('files[]')
    is_sales = request.form.get('is_sales', 'false').lower() == 'true'
    
    results = []
    for f in files:
        if f.filename == '':
            continue
            
        # Save file to temp directory
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], f.filename)
        f.save(temp_path)
        
        try:
            # Parse the invoice file
            data = process_file(temp_path, is_sales)
            results.append({
                'filename': f.filename,
                'status': 'success',
                'data': data
            })
        except Exception as e:
            results.append({
                'filename': f.filename,
                'status': 'error',
                'error': str(e)
            })
        finally:
            # Clean up temp file
            if os.path.exists(temp_path):
                os.remove(temp_path)
                
    return jsonify({'results': results})

@app.route('/api/export', methods=['POST'])
def export_excel():
    data = request.json or {}
    purchase_list = data.get('purchase', [])
    sales_list = data.get('sales', [])
    report_month = data.get('month', 'May 2026')
    
    wb = openpyxl.Workbook()
    wb.properties.fullCalcOnLoad = True
    ws = wb.active
    ws.title = "GST Summary"
    ws.views.sheetView[0].showGridLines = True
    
    # Common Styles
    title_font = Font(name='Segoe UI', size=14, bold=True, color='2C3E50')
    section_font = Font(name='Segoe UI', size=11, bold=True, color='2C3E50')
    header_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    bold_font = Font(name='Segoe UI', size=10, bold=True)
    regular_font = Font(name='Segoe UI', size=10)
    
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    section_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    total_fill = PatternFill(start_color='E4ECF5', end_color='E4ECF5', fill_type='solid')
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    double_side = Side(border_style="double", color="000000")
    thick_bottom = Side(border_style="medium", color="000000")
    
    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_total = Border(top=thin_side, bottom=double_side)
    border_net = Border(top=thin_side, bottom=thick_bottom)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # 1. Main Title
    ws.merge_cells('A1:K1')
    ws['A1'] = f"GST summary for the month of {report_month}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 30
    
    current_row = 3
    
    # --- Function to write a table ---
    def write_table(title, items, is_sale_table=False):
        nonlocal current_row
        
        # Section Header
        ws.cell(row=current_row, column=1, value=title).font = section_font
        ws.cell(row=current_row, column=1).alignment = align_left
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        
        # Column Headers
        headers = [
            "Sr No", "Date", "Name of Party", "GST Number", "Invoice Number", 
            "GST Rate", "Taxable Value (Rs.)", "CGST (Rs.)", "SGST (Rs.)", 
            "IGST (Rs.)", "Total Invoice Value (Rs.)"
        ]
        
        ws.row_dimensions[current_row].height = 24
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_cell
            
        header_row_idx = current_row
        current_row += 1
        
        start_data_row = current_row
        
        # Data Rows
        for idx, item in enumerate(items, 1):
            ws.row_dimensions[current_row].height = 20
            
            # Format rate values
            rate_val = item.get('gst_rate', 0.0)
            taxable_val = float(item.get('taxable_value', 0.0))
            cgst_val = float(item.get('cgst', 0.0))
            sgst_val = float(item.get('sgst', 0.0))
            igst_val = float(item.get('igst', 0.0))
            total_val = float(item.get('total_invoice_value', 0.0))
            
            # Cells
            c_sr = ws.cell(row=current_row, column=1, value=idx)
            c_date = ws.cell(row=current_row, column=2, value=item.get('date', ''))
            c_party = ws.cell(row=current_row, column=3, value=item.get('party_name', ''))
            c_gst = ws.cell(row=current_row, column=4, value=item.get('gst_number', ''))
            c_inv = ws.cell(row=current_row, column=5, value=item.get('invoice_number', ''))
            c_rate = ws.cell(row=current_row, column=6, value=rate_val)
            c_tax = ws.cell(row=current_row, column=7, value=taxable_val)
            c_cgst = ws.cell(row=current_row, column=8, value=cgst_val)
            c_sgst = ws.cell(row=current_row, column=9, value=sgst_val)
            c_igst = ws.cell(row=current_row, column=10, value=igst_val)
            c_tot = ws.cell(row=current_row, column=11, value=total_val)
            
            # Formats and Fonts
            for c in [c_sr, c_date, c_party, c_gst, c_inv, c_rate, c_tax, c_cgst, c_sgst, c_igst, c_tot]:
                c.font = regular_font
                c.border = border_cell
                
            c_sr.alignment = align_center
            c_date.alignment = align_center
            c_party.alignment = align_left
            c_gst.alignment = align_center
            c_inv.alignment = align_center
            
            c_rate.number_format = '0%'
            c_rate.alignment = align_center
            
            for amt_cell in [c_tax, c_cgst, c_sgst, c_igst, c_tot]:
                amt_cell.number_format = '#,##0.00'
                amt_cell.alignment = align_right
                
            current_row += 1
            
        end_data_row = current_row - 1
        
        # Total Row
        ws.row_dimensions[current_row].height = 22
        
        # Merge/label for Total
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        total_label = ws.cell(row=current_row, column=1, value="Total")
        total_label.font = bold_font
        total_label.alignment = align_right
        
        # Apply borders to merged total cells
        for c_idx in range(1, 6):
            ws.cell(row=current_row, column=c_idx).border = border_total
            ws.cell(row=current_row, column=c_idx).fill = total_fill
            
        # Empty Rate Cell
        ws.cell(row=current_row, column=6).border = border_total
        ws.cell(row=current_row, column=6).fill = total_fill
        
        # Sum Formulas
        tax_total_cell = ws.cell(row=current_row, column=7)
        cgst_total_cell = ws.cell(row=current_row, column=8)
        sgst_total_cell = ws.cell(row=current_row, column=9)
        igst_total_cell = ws.cell(row=current_row, column=10)
        tot_total_cell = ws.cell(row=current_row, column=11)
        
        if len(items) > 0:
            tax_total_cell.value = f"=SUM(G{start_data_row}:G{end_data_row})"
            cgst_total_cell.value = f"=SUM(H{start_data_row}:H{end_data_row})"
            sgst_total_cell.value = f"=SUM(I{start_data_row}:I{end_data_row})"
            igst_total_cell.value = f"=SUM(J{start_data_row}:J{end_data_row})"
            tot_total_cell.value = f"=SUM(K{start_data_row}:K{end_data_row})"
        else:
            tax_total_cell.value = 0.0
            cgst_total_cell.value = 0.0
            sgst_total_cell.value = 0.0
            igst_total_cell.value = 0.0
            tot_total_cell.value = 0.0
            
        for c in [tax_total_cell, cgst_total_cell, sgst_total_cell, igst_total_cell, tot_total_cell]:
            c.font = bold_font
            c.border = border_total
            c.fill = total_fill
            c.number_format = '#,##0.00'
            c.alignment = align_right
            
        total_row_idx = current_row
        current_row += 2 # Leave space after table
        
        return total_row_idx
        
    # Write Purchase Table
    purchase_total_row = write_table("Purchase", purchase_list, False)
    
    # Write Sales Table
    sales_total_row = write_table("Sale", sales_list, True)
    
    # 2. Net Payable Table (Summary)
    ws.cell(row=current_row, column=1, value="Net Summary").font = section_font
    ws.cell(row=current_row, column=1).alignment = align_left
    ws.row_dimensions[current_row].height = 22
    current_row += 1
    
    ws.row_dimensions[current_row].height = 22
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    net_label = ws.cell(row=current_row, column=1, value="Net Payable")
    net_label.font = bold_font
    net_label.alignment = align_right
    
    # Fill background and border for merged cells
    for c_idx in range(1, 6):
        ws.cell(row=current_row, column=c_idx).border = border_net
        ws.cell(row=current_row, column=c_idx).fill = section_fill
        
    # Empty rate cell
    ws.cell(row=current_row, column=6).border = border_net
    ws.cell(row=current_row, column=6).fill = section_fill
    
    # Net values: Sales total minus Purchase total using formulas
    net_tax_cell = ws.cell(row=current_row, column=7, value=f"=G{sales_total_row}-G{purchase_total_row}")
    net_cgst_cell = ws.cell(row=current_row, column=8, value=f"=H{sales_total_row}-H{purchase_total_row}")
    net_sgst_cell = ws.cell(row=current_row, column=9, value=f"=I{sales_total_row}-I{purchase_total_row}")
    net_igst_cell = ws.cell(row=current_row, column=10, value=f"=J{sales_total_row}-J{purchase_total_row}")
    net_tot_cell = ws.cell(row=current_row, column=11, value=f"=K{sales_total_row}-K{purchase_total_row}")
    
    for c in [net_tax_cell, net_cgst_cell, net_sgst_cell, net_igst_cell, net_tot_cell]:
        c.font = bold_font
        c.border = border_net
        c.fill = section_fill
        c.number_format = '#,##0.00'
        c.alignment = align_right
        
    # Set generous default column widths to prevent truncation and ### display errors
    col_widths = {
        'A': 8,   # Sr No
        'B': 14,  # Date
        'C': 35,  # Name of Party
        'D': 18,  # GST Number
        'E': 18,  # Invoice Number
        'F': 12,  # GST Rate
        'G': 22,  # Taxable Value (Rs.)
        'H': 18,  # CGST (Rs.)
        'I': 18,  # SGST (Rs.)
        'J': 18,  # IGST (Rs.)
        'K': 22   # Total Invoice Value (Rs.)
    }
    
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Save to temp file and return
    temp_excel = os.path.join(app.config['UPLOAD_FOLDER'], "GST_Summary_Report.xlsx")
    wb.save(temp_excel)
    
    return send_file(temp_excel, as_attachment=True, download_name="GST_Summary_Report.xlsx")

if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5000, debug=True)
